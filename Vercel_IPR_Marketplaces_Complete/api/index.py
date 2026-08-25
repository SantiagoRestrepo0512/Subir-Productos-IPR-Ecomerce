from flask import Flask, request, send_file, jsonify, render_template_string
import io, re, html, os, glob
import openpyxl

app = Flask(__name__)

# Función para encontrar automáticamente el archivo CSV dentro de la carpeta api/
def get_csv_path():
    base_dir = os.path.dirname(__file__)
    csv_files = glob.glob(os.path.join(base_dir, "*.csv"))
    if csv_files:
        return csv_files[0]
    return None

BTU_SET = {9000, 12000, 18000, 24000, 30000, 36000, 42000, 48000, 50000, 60000}

LISTA_MARCAS_CONOCIDAS = [
    "MIDEA", "SAMSUNG", "LG", "CARRIER", "TRANE", "YORK", "DAIKIN", 
    "COPELAND", "DANFOSS", "EMERSON", "PANASONIC", "GREE", "HACEB", 
    "HISENSE", "ELECTROLUX", "GENERAL ELECTRIC", "HONEYWELL", "MABE",
    "BOSCH", "WESTINGHOUSE", "WHITE RODGERS", "ELGIN", "EMBRACO",
    "TECUMSEH", "FULL GAUGE", "DIXELL", "SOLER Y PALAU", "ROWA",
    "ASPEN", "ELITECH", "DORIN", "MCQUAY", "DUPONT", "CHEMOURS",
    "3M", "LOCTITE", "TYCO", "WEG", "CHEVRON"
]

MARCAS_ML = {
    "BOSCH": "BOSCH", "CARRIER": "CARRIER", "CHEVRON": "CHEVRON", "COPELAND": "COPELAND",
    "DAIKIN": "DAIKIN", "DANFOSS": "DANFOSS", "DUPONT": "DUPONT", "ELECTROLUX": "ELECTROLUX",
    "EMERSON": "EMERSON", "GENERAL ELECTRIC": "GENERAL ELECTRIC", "GREE": "GREE",
    "HACEB": "HACEB", "HISENSE": "HISENSE", "HONEYWELL": "HONEYWELL", "LG": "LG",
    "LOCTITE": "LOCTITE", "MABE": "MABE", "MIDEA": "MIDEA", "PANASONIC": "PANASONIC",
    "SAMSUNG": "SAMSUNG", "TRANE": "TRANE", "TYCO": "TYCO", "WEG": "WEG",
    "WESTINGHOUSE": "WESTINGHOUSE", "WHITE RODGERS": "WHITE RODGERS", "3M": "3M"
}

MARCAS_FALABELLA = {
    "BOSCH": "BOSCH", "CARRIER": "CARRIER", "CHEVRON": "CHEVRON", "COPELAND": "COPELAND",
    "DAIKIN": "DAIKIN", "DANFOSS": "DANFOSS", "DUPONT": "DUPONT", "ELECTROLUX": "ELECTROLUX",
    "EMERSON": "EMERSON", "GENERAL ELECTRIC": "GENERAL ELECTRIC", "GREE": "GENERICO",
    "HACEB": "HACEB", "HISENSE": "HISENSE", "HONEYWELL": "HONEYWELL", "LG": "LG",
    "LOCTITE": "LOCTITE", "MABE": "MABE", "MIDEA": "MIDEA", "PANASONIC": "PANASONIC",
    "SAMSUNG": "SAMSUNG", "TRANE": "TRANE", "TYCO": "TYCO", "WEG": "WEG",
    "WESTINGHOUSE": "WESTINGHOUSE", "WHITE RODGERS": "WHITE RODGERS", "3M": "3M",
    "ABRO": "GENERICO", "AKO": "GENERICO", "ALADO": "GENERICO", "ALFA FLUX": "GENERICO",
    "ALFA REFRIGERACION": "GENERICO", "AMUCO": "GENERICO", "ASPEN": "GENERICO", "ASPEN PUMP": "GENERICO",
    "BALFLEX": "GENERICO", "BERNZOMATIC": "GENERICO", "BNF": "GENERICO", "BREAKERMATIC": "GENERICO",
    "BRILLA AL-CO": "GENERICO", "BRISTOL": "GENERICO", "BUILDERS BEST": "GENERICO", "CARLYLE": "GENERICO",
    "CELLUX": "GENERICO", "CHEMOURS": "GENERICO", "CIAC": "GENERICO", "CLEARWAY": "GENERICO",
    "COMFORT AIR": "GENERICO", "COMFORT STAR": "GENERICO", "COMFORT STYLE": "GENERICO", "COMPELA": "GENERICO",
    "COMPTEK": "GENERICO", "CONALCABLES": "GENERICO", "COPPER TUBE": "GENERICO", "CUBIGEL": "GENERICO",
    "DELTA FRÃO": "GENERICO", "DELTA FRIO": "GENERICO", "DELTAFRIO": "GENERICO", "DIVERSITECH": "GENERICO",
    "DIXELL": "GENERICO", "DORIN": "GENERICO", "DWYER": "GENERICO", "DYNAIR": "GENERICO",
    "ELCO": "GENERICO", "ELGIN": "GENERICO", "ELITECH": "GENERICO", "EMBRACO": "GENERICO",
    "EMICOL": "GENERICO", "ERRECOM": "GENERICO", "ESPUMLATEX": "GENERICO", "EUROPAN": "GENERICO",
    "EVERWELL": "GENERICO", "FASCO": "GENERICO", "FIBERGLASS ISO": "GENERICO", "FSP": "GENERICO",
    "FULL GAUGE": "GENERICO", "GBP": "GENERICO", "GENÉRICO": "GENERICO", "GENA©RICO": "GENERICO",
    "GENPRO": "GENERICO", "GMCC": "GENERICO", "HAMMER": "GENERICO", "HARRIS": "GENERICO",
    "HARTLAND": "GENERICO", "HUAYI": "GENERICO", "INCOPAR": "GENERICO", "INNOVACION C": "GENERICO",
    "INTERKLIMAT": "GENERICO", "INVENSYS": "GENERICO", "INVOTECH": "GENERICO", "IPR": "GENERICO",
    "JASON": "GENERICO", "JOHNSON CONTROLS": "GENERICO", "KINGSPAN": "GENERICO", "KLIXON": "GENERICO",
    "KOCH GREEN": "GENERICO", "LACO": "GENERICO", "LONG TERM": "GENERICO", "MADEPREN": "GENERICO",
    "MANEUROP": "GENERICO", "MARS": "GENERICO", "MASTERCOOL": "GENERICO", "MAXPURE": "GENERICO",
    "MCQUAY": "GENERICO", "MINEROIL": "GENERICO", "MIPAL": "GENERICO", "MIRAGE": "GENERICO",
    "NACOBRE": "GENERICO", "NINGBO": "GENERICO", "ORANGE": "GENERICO", "PARAGON": "GENERICO",
    "PEGAUCHO": "GENERICO", "PENN": "GENERICO", "POLYKEN": "GENERICO", "POLYLON": "GENERICO",
    "RANCO": "GENERICO", "READY": "GENERICO", "REFRIANDINOS": "GENERICO", "RIFENG": "GENERICO",
    "ROBERTSHAW": "GENERICO", "ROWA": "GENERICO", "ROYALSTAR": "GENERICO", "SANHUA": "GENERICO",
    "SICCOM": "GENERICO", "SINTECO": "GENERICO", "SMART ELECTRIC": "GENERICO", "SMART PUMP": "GENERICO",
    "SOLER Y PALAU": "GENERICO", "SUMOIL": "GENERICO", "SUPCO": "GENERICO", "TECAM": "GENERICO",
    "TECLAB": "GENERICO", "TECNOWELD": "GENERICO", "TECUMSEH": "GENERICO", "THERMO-COIL": "GENERICO",
    "TIANYICOOL": "GENERICO", "TOP TECH": "GENERICO", "TOPFLO": "GENERICO", "TRADEPRO": "GENERICO",
    "UNCO": "GENERICO", "UNIWELD": "GENERICO", "US MOTORS": "GENERICO", "WAGNER": "GENERICO",
    "YETI": "GENERICO", "ZERO": "GENERICO", "ZIEHL-ABEGG": "GENERICO", "ZI ZHENG LAN": "GENERICO"
}

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Hub de Marketplaces IPR</title>
  <style>
    :root {
      --primary: #2563eb; --bg: #0b0f19; --card: #161f30; --card-inner: #0f172a;
      --text: #f8fafc; --text-muted: #94a3b8; --border: #334155; --success: #10b981;
    }
    * { box-sizing: border-box; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
    body { background: var(--bg); color: var(--text); padding: 30px 15px; display: flex; justify-content: center; margin: 0; }
    .container { width: 100%; max-width: 680px; background: var(--card); border-radius: 16px; border: 1px solid var(--border); box-shadow: 0 20px 40px rgba(0,0,0,0.5); padding: 35px; }
    .header { text-align: center; margin-bottom: 25px; }
    .header h1 { font-size: 1.8rem; margin: 0 0 6px 0; color: #ffffff; }
    .header p { color: var(--text-muted); margin: 0; font-size: 0.95rem; }
    .market-tabs { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; margin-bottom: 25px; }
    .tab-btn {
      background: var(--card-inner); border: 2px solid var(--border); color: var(--text-muted); padding: 14px 10px;
      border-radius: 12px; cursor: pointer; text-align: center; font-weight: 600; font-size: 0.95rem;
      transition: all 0.2s ease; user-select: none;
    }
    .tab-btn:hover { border-color: #64748b; color: #ffffff; }
    .tab-btn.active { border-color: #38bdf8; background: rgba(56, 189, 248, 0.12); color: #38bdf8; }
    .form-group { margin-bottom: 20px; }
    label { display: block; font-weight: 600; margin-bottom: 8px; font-size: 0.9rem; color: #cbd5e1; }
    input[type="file"], textarea, input[type="number"] {
      width: 100%; padding: 12px 14px; border: 1px solid var(--border); border-radius: 8px;
      font-size: 0.95rem; background: var(--card-inner); color: #ffffff;
    }
    input[type="file"]::file-selector-button {
      background: #334155; color: white; border: none; padding: 6px 12px; border-radius: 6px; cursor: pointer; margin-right: 10px;
    }
    textarea { height: 120px; resize: vertical; font-family: monospace; font-size: 0.9rem; }
    .badge-info {
      display: inline-block; background: rgba(16, 185, 129, 0.15); color: #34d399; font-size: 0.8rem;
      padding: 5px 10px; border-radius: 6px; margin-top: 6px;
    }
    button.btn-submit {
      width: 100%; background: var(--success); color: white; border: none; padding: 16px; border-radius: 10px;
      font-size: 1.05rem; font-weight: bold; cursor: pointer; transition: background 0.2s; margin-top: 10px;
    }
    button.btn-submit:hover { background: #059669; }
    button.btn-submit:disabled { background: #475569; cursor: not-allowed; }
    #status { margin-top: 20px; font-size: 0.95rem; text-align: center; font-weight: 500; }
  </style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>📦 Hub de Marketplaces IPR</h1>
    <p>Catálogo maestro cargado en el servidor</p>
  </div>

  <form id="unifiedForm">
    <input type="hidden" id="marketplaceInput" name="marketplace" value="ml">

    <label>1. Selecciona el Marketplace destino:</label>
    <div class="market-tabs">
      <div class="tab-btn active" onclick="selectMarket('ml', 22, 'Mercado Libre')">🟡 Mercado Libre</div>
      <div class="tab-btn" onclick="selectMarket('falabella', 17, 'Falabella')">🟢 Falabella</div>
      <div class="tab-btn" onclick="selectMarket('exito', 22, 'Éxito')">🔴 Éxito</div>
    </div>

    <div class="form-group">
      <label id="lblPlantilla">2. Plantilla Excel Mercado Libre (.xlsx):</label>
      <input type="file" id="xlsxFile" name="xlsx_file" accept=".xlsx" required>
    </div>

    <div class="form-group">
      <label>Margen de Ganancia (%):</label>
      <input type="number" id="margen" name="margen" value="22" min="0" max="200" step="0.5">
      <div class="badge-info" id="badgeMargen">Margen predeterminado: 22%</div>
    </div>

    <div class="form-group">
      <label>3. SKUs a procesar (uno por línea o separados por coma):</label>
      <textarea id="skus" name="skus" placeholder="614-0109&#10;614-0108&#10;614-0107" required>614-0109
614-0108
614-0107</textarea>
    </div>

    <button type="submit" class="btn-submit" id="btnSubmit">⚡ GENERAR EXCEL PARA MERCADO LIBRE</button>
  </form>

  <div id="status"></div>
</div>

<script>
let currentMarket = 'ml';
function selectMarket(market, defaultMargin, marketName) {
  currentMarket = market;
  document.getElementById('marketplaceInput').value = market;
  document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
  event.currentTarget.classList.add('active');
  document.getElementById('margen').value = defaultMargin;
  document.getElementById('badgeMargen').textContent = `Margen predeterminado para ${marketName}: ${defaultMargin}%`;
  document.getElementById('lblPlantilla').textContent = `2. Plantilla Excel ${marketName} (.xlsx):`;
  document.getElementById('btnSubmit').textContent = `⚡ GENERAR EXCEL PARA ${marketName.toUpperCase()}`;
}

document.getElementById('unifiedForm').addEventListener('submit', async (e) => {
  e.preventDefault();
  const btn = document.getElementById('btnSubmit');
  const status = document.getElementById('status');
  btn.disabled = true;
  btn.textContent = "⏳ Procesando en Vercel...";
  status.textContent = "Buscando productos en el catálogo maestro y generando plantilla...";
  status.style.color = "#38bdf8";

  const formData = new FormData(document.getElementById('unifiedForm'));

  try {
    const res = await fetch('/api/process', { method: 'POST', body: formData });
    if (!res.ok) {
      const errData = await res.json().catch(() => ({}));
      throw new Error(errData.error || `Error en el servidor (${res.status})`);
    }

    const blob = await res.blob();
    const url = window.URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    
    let defaultNames = {
      'ml': 'mercadolibre_ipr_FILTRADO_FINAL.xlsx',
      'falabella': 'fallabela_ipr_FILTRADO_FINAL.xlsx',
      'exito': 'exito_ipr_FILTRADO_FINAL_ESTANDARIZADO.xlsx'
    };
    
    a.download = defaultNames[currentMarket] || "catalogo_procesado.xlsx";
    document.body.appendChild(a);
    a.click();
    a.remove();
    window.URL.revokeObjectURL(url);

    status.textContent = `✅ ¡Archivo ${a.download} generado y descargado con éxito!`;
    status.style.color = "#34d399";
  } catch (err) {
    status.textContent = "❌ Error: " + err.message;
    status.style.color = "#f87171";
    alert("Ocurrió un error: " + err.message);
  } finally {
    btn.disabled = false;
    btn.textContent = `⚡ GENERAR EXCEL PARA ${currentMarket.toUpperCase()}`;
  }
});
</script>
</body>
</html>
"""

def extract_id(block):
    m = re.match(r'"(\d+),simple,', block)
    return int(m.group(1)) if m else None

def extract_precio(block):
    rid = extract_id(block)
    cand = []
    for n in re.findall(r'\b\d{5,8}\b', block):
        v = int(n)
        if v in BTU_SET or 1900 <= v <= 2099 or (rid is not None and v == rid):
            continue
        cand.append(v)
    if cand: return max(cand)
    m = re.search(r'(\d+),producto/', block)
    if m: return int(m.group(1))
    m = re.search(r'(\d+),""Refrigeraci', block)
    if m: return int(m.group(1))
    return 0

def extract_nombre(block, sku):
    m = re.search(r'simple,' + re.escape(sku) + r',,""([^"](?:[^"]|"")*?)"",', block)
    return m.group(1).replace('""', '"') if m else ""

def limpiar_nombre(n):
    n = re.sub(r'[^\w\s]', ' ', n, flags=re.UNICODE).replace('_', ' ')
    return re.sub(r'\s+', ' ', n).strip()

def extract_desc(block):
    m = re.search(r'visible,(.*?)(?:,,,taxable|,producto/|,""Refrigeraci|,,,\d{4,8},|$)', block, re.S)
    region = m.group(1) if m else ''
    prev = ""
    while prev != region:
        prev = region
        region = html.unescape(region)
    region = region.replace('\\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    region = re.sub(r'<(script|style)[^>]*>.*?</\1>', ' ', region, flags=re.I | re.S)
    region = re.sub(r'<[^>]+>', ' ', region)
    text = region.replace('""', '"').replace('"', ' ')
    text = re.sub(r'[;\s]+', ' ', text).strip()
    text = re.sub(r'^[A-Z0-9/.\-]+\s*,\s*', '', text)
    return re.sub(r'^[^\w]+', '', text).strip()

def extract_marca_raw(block):
    m = re.search(r'producto/[^,]*,\s*([^,]+),', block)
    if m and m.group(1).strip().lower() not in ('1', '0', 'none', ''):
        return m.group(1).strip()
    m2 = re.search(r'"Refrigeraci.*?"",""([^"]+)"",', block)
    if m2 and m2.group(1).strip().lower() not in ('1', '0', 'none', ''):
        return m2.group(1).strip()
    return ""

def procesar_marca_exito(marca_raw, texto=""):
    m = marca_raw.upper().strip()
    if m and m not in ('GENÉRICO', 'GENERICO', 'GENA©RICO', 'NONE', '0', '1', 'N/A', 'SIN MARCA'):
        return m
    texto_upper = texto.upper()
    for marca in LISTA_MARCAS_CONOCIDAS:
        if re.search(r'\b' + re.escape(marca) + r'\b', texto_upper):
            return marca
    return "GENERICO"

def extract_imgs(block):
    u = re.findall(r'https?://[^\s"<>,]+\.(?:jpg|jpeg|png|gif|webp)', block, re.I)
    seen = []
    for x in u:
        if x not in seen: seen.append(x)
    return seen

def generar_ean13(sku):
    digits = re.sub(r'\D', '', str(sku))
    base12 = f"770{digits.zfill(9)}"[-12:]
    s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(base12))
    check_digit = (10 - (s % 10)) % 10
    return base12 + str(check_digit)

def deducir_capacidad_ml(text):
    m = re.search(r'(\d{1,2}(?:[\.,]\d{3})?|\d{4,5})\s*BTU', text, re.I)
    if m:
        num = int(re.sub(r'\D', '', m.group(1)))
        valid_btus = [5000, 9000, 11000, 12000, 14000, 17000, 18000, 22000, 24000, 30000, 36000, 48000, 60000]
        return min(valid_btus, key=lambda x: abs(x - num))
    return 9000

def deducir_voltaje_ml(text):
    m = re.search(r'(\d{3})\s*V', text, re.I)
    if m:
        v = int(m.group(1))
        if v in [110, 115, 120, 220, 230, 240]:
            return "110V" if v < 130 else "220V"
    return "220V"

def deducir_tipo_alimentacion(text):
    if re.search(r'trif[aá]sico|380\s*v|440\s*v|460\s*v|480\s*v|industrial', text, re.I):
        return "Energía industrial"
    return "Corriente doméstica"

def deducir_refrigerante_ml(text):
    for r in ['R-410A', 'R-32', 'R-22', 'R134', 'R600A']:
        if re.search(r'\b' + re.escape(r) + r'\b', text, re.I): return r
    return "R-410A"

def deducir_climatizacion(text):
    return "Frío/Calor" if re.search(r'calor|calefacci[oó]n', text, re.I) else "Frío"

def deducir_tipo_aire(text):
    if re.search(r'portatil|portátil', text, re.I): return "Portátil"
    if re.search(r'ventana', text, re.I): return "Ventana"
    return "Split"

def deducir_capacidad_falabella(t):
    m = re.search(r'(\d[\d .]{2,6})\s*BTU', t, re.I)
    if m: return (re.sub(r'\D', '', m.group(1)) + ' BTU').strip()
    m = re.search(r'(\d{1,2})\s*K\b', t)
    if m: return m.group(0).strip()
    m = re.search(r'(\d+(?:\.\d+)?)\s*HP', t, re.I)
    if m: return m.group(0).strip()
    return ''

def deducir_tension_falabella(t):
    m = re.search(r'(\d{3})\s*V', t, re.I)
    return m.group(1) if m else ''

def deducir_capacidad_exito(text):
    m = re.search(r'(\d{1,2}(?:[\.,]\d{3})?|\d{4,5})\s*BTU', text, re.I)
    if m:
        num = int(re.sub(r'\D', '', m.group(1)))
        valid_btus = [5000, 9000, 11000, 12000, 14000, 17000, 18000, 22000, 24000]
        closest = min(valid_btus, key=lambda x: abs(x - num))
        return f"{closest} Btu"
    return "9000 Btu"

def deducir_voltaje_exito(text):
    m = re.search(r'(\d{3})\s*V', text, re.I)
    if m:
        v = int(m.group(1))
        if v in [110, 115, 120, 220, 230, 240]:
            return f"{v} V"
    return "220 V"

def deducir_refrigerante_exito(text):
    for r in ['R-410A', 'R-32', 'R-22', 'R134', 'R600A']:
        if re.search(r'\b' + re.escape(r) + r'\b', text, re.I):
            return r
    return "Otro"

def deducir_tecnologia_exito(text):
    if re.search(r'dual\s+inverter', text, re.I): return "Dual Inverter"
    if re.search(r'inverter', text, re.I): return "Inverter"
    return "Convencional"

@app.route('/', methods=['GET'])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/process', methods=['POST'])
def process():
    try:
        csv_path = get_csv_path()
        if not csv_path or not os.path.exists(csv_path):
            return jsonify({"error": "No se encontró ningún archivo .csv en la carpeta api/"}), 500

        marketplace = request.form.get('marketplace', 'ml')
        skus_raw = request.form.get('skus', '')
        margen_pct = float(request.form.get('margen', 22))
        factor_margen = 1.0 + (margen_pct / 100.0)

        xlsx_file = request.files.get('xlsx_file')
        if not xlsx_file:
            return jsonify({"error": "Debes adjuntar la plantilla Excel oficial."}), 400

        with open(csv_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            raw = f.read()

        starts = [m.start() for m in re.finditer(r'(?m)^"\d+', raw)]
        starts.append(len(raw))

        def get_block(sku):
            pos = raw.find(sku)
            if pos < 0: return None
            for i, st in enumerate(starts[:-1]):
                if st <= pos < starts[i+1]:
                    return raw[st:starts[i+1]]
            return None

        skus_list = [s.strip() for s in re.split(r'[\n,]+', skus_raw) if s.strip()]

        parsed_items = []
        for sku in skus_list:
            b = get_block(sku)
            if not b: continue
            nombre = limpiar_nombre(extract_nombre(b, sku))
            precio = extract_precio(b)
            desc = extract_desc(b) or "Sin descripción"
            marca_raw = extract_marca_raw(b)
            imgs = extract_imgs(b)
            texto = f"{nombre} {desc}"

            parsed_items.append({
                "sku": sku, "nombre": nombre, "precio": precio, "desc": desc,
                "marca_raw": marca_raw, "imgs": imgs, "texto": texto
            })

        if not parsed_items:
            return jsonify({"error": "Ninguno de los SKUs fue encontrado en el catálogo maestro."}), 400

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()))

        if marketplace == "ml":
            ws = wb['Aires Acondicionados'] if 'Aires Acondicionados' in wb.sheetnames else wb.sheetnames[0]
            for r in range(ws.max_row, 8, -1): ws.delete_rows(r)
            row = 9
            for p in parsed_items:
                precio_final = round(p['precio'] * factor_margen)
                
                # Rescate de marca para ML
                marca_ml = MARCAS_ML.get(p['marca_raw'].upper().strip())
                if not marca_ml:
                    for km in MARCAS_ML.keys():
                        if re.search(r'\b' + re.escape(km) + r'\b', p['texto'].upper()):
                            marca_ml = MARCAS_ML[km]
                            break
                marca_ml = marca_ml or "Genérica"

                tipo_aire = deducir_tipo_aire(p['texto'])
                row_data = {
                    2: p['nombre'][:60], 4: "Nuevo", 5: generar_ean13(p['sku']), 6: "Blanco",
                    7: deducir_voltaje_ml(p['texto']), 8: deducir_voltaje_ml(p['texto']),
                    9: ",".join(p['imgs'][:10]), 10: p['sku'], 11: 1, 12: precio_final, 13: p['desc'],
                    15: "Cuotas", 16: "Sin costo", 17: "Mercado Envíos", 18: "A cargo del comprador",
                    19: "Acepto", 20: "Garantía del vendedor", 21: 12, 22: "meses", 23: marca_ml,
                    24: p['sku'], 25: deducir_tipo_alimentacion(p['texto']), 26: tipo_aire,
                    28: deducir_climatizacion(p['texto']), 29: "De pared" if tipo_aire == "Split" else "",
                    34: "Sí", 35: deducir_capacidad_ml(p['texto']), 36: "BTU",
                    37: "Sí" if "wifi" in p['texto'].lower() or "wi-fi" in p['texto'].lower() else "No",
                    38: "Sí", 39: "Sí", 40: deducir_refrigerante_ml(p['texto']),
                    41: "Sí" if "inverter" in p['texto'].lower() else "No"
                }
                for c, val in row_data.items(): ws.cell(row=row, column=c, value=val)
                row += 1
            out_name = "mercadolibre_ipr_FILTRADO_FINAL.xlsx"

        elif marketplace == "falabella":
            ws = wb['Subir plantilla'] if 'Subir plantilla' in wb.sheetnames else wb.sheetnames[0]
            for r in range(ws.max_row, 4, -1): ws.delete_rows(r)
            row = 5
            for p in parsed_items:
                precio_final = round(p['precio'] * factor_margen)
                
                # Rescate de marca para Falabella
                marca_fal = MARCAS_FALABELLA.get(p['marca_raw'].upper().strip())
                if not marca_fal or marca_fal == "GENERICO":
                    for km in MARCAS_FALABELLA.keys():
                        if re.search(r'\b' + re.escape(km) + r'\b', p['texto'].upper()):
                            marca_fal = MARCAS_FALABELLA[km]
                            break
                marca_fal = marca_fal or "GENERICO"

                vals = {
                    1: p['nombre'], 2: marca_fal, 3: p['nombre'], 4: p['desc'], 5: 2376,
                    7: p['sku'], 8: p['sku'].replace('-', ''), 9: 'Sin variación', 10: 19, 11: 1,
                    12: precio_final, 16: deducir_tension_falabella(p['texto']), 17: deducir_capacidad_falabella(p['texto']),
                    18: 0, 19: 0, 20: 'Unidad', 41: 'Nuevo', 47: 100, 48: 40, 49: 35, 50: 40
                }
                for c, v in vals.items(): ws.cell(row=row, column=c, value=v)
                for j, u in enumerate(p['imgs'][:8]): ws.cell(row=row, column=51 + j, value=u)
                row += 1
            out_name = "fallabela_ipr_FILTRADO_FINAL.xlsx"

        elif marketplace == "exito":
            ws = wb['Aires Acondicionados'] if 'Aires Acondicionados' in wb.sheetnames else wb.sheetnames[0]
            for r in range(ws.max_row, 3, -1): ws.delete_rows(r)
            row = 4
            for p in parsed_items:
                marca = procesar_marca_exito(p['marca_raw'], p['texto'])
                imgs = p['imgs']
                row_data = {
                    1: "", 2: p['sku'].replace('-', ''), 3: p['nombre'][:120], 4: "27432_Aires Acondicionados",
                    5: marca, 6: p['desc'], 7: f"{marca}, {p['sku']}, Aire Acondicionado, Refrigeracion",
                    8: 35, 9: 100, 10: 40, 11: 40, 12: 3, 13: 35, 14: 100, 15: 40, 16: 40, 17: "Unidad",
                    18: 1, 19: "Technology", 20: imgs[0] if len(imgs) > 0 else "N/A",
                    21: imgs[1] if len(imgs) > 1 else "", 22: imgs[2] if len(imgs) > 2 else "",
                    23: imgs[3] if len(imgs) > 3 else "", 24: imgs[4] if len(imgs) > 4 else "",
                    25: "", 26: 0, 27: "N/A", 28: "N/A", 29: "Sin advertencias específicas",
                    30: deducir_tipo_aire(p['texto']), 31: deducir_capacidad_exito(p['texto']),
                    32: deducir_voltaje_exito(p['texto']), 33: deducir_refrigerante_exito(p['texto']),
                    34: "N/A", 35: "Control remoto, Manual de usuario", 36: "N/A",
                    37: deducir_tecnologia_exito(p['texto']), 38: "No Aplica"
                }
                for c, val in row_data.items(): ws.cell(row=row, column=c, value=val)
                row += 1
            out_name = "exito_ipr_FILTRADO_FINAL_ESTANDARIZADO.xlsx"

        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)

        return send_file(
            out_buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000)
